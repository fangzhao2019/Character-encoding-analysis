# coding:utf-8
from pymongo import MongoClient
from openpyxl import Workbook

client=MongoClient('192.168.1.94',30000)
collection=client.car_autohome.ttf_manage

wb=Workbook()
ws=wb.active
ws.cell(row=1,column=1).value='idx'
ws.cell(row=1,column=2).value='ttf_name'
ws.cell(row=1,column=3).value='code'
ws.cell(row=1,column=4).value='word'

idx=1
for row in collection.find({}):
    ttf_name=row['ttf']
    data=row['data']
    for key in data.keys():
        idx+=1
        if idx%1000==0:print(idx)
        word=data[key]
        ws.cell(row=idx,column=1).value=idx-1
        ws.cell(row=idx,column=2).value=ttf_name
        ws.cell(row=idx,column=3).value=key
        try:
            len(word)>0
            ws.cell(row=idx,column=4).value=word
        except:
            ws.cell(row=idx,column=4).value='null'
            

wb.save('init_coding.xlsx')
            
