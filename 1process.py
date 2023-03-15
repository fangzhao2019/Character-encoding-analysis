import re
from openpyxl import load_workbook
from openpyxl import Workbook
from tqdm import tqdm
import math
import operator
import json
import jieba
import jieba.posseg as pseg
jieba.load_userdict('../userdict.txt')


# 统计词频，并返回字典
def freq_word(word_list):
    fre_word = {}
    for w in word_list:
        if str(w) in fre_word:
            fre_word[str(w)] += 1
        else:
            fre_word[str(w)] = 1
    return fre_word

# 查出包含每个词的文档数
def word_in_file_count(article_list):
    word_file_count = {"total article num": len(article_list)}
    for article in article_list:
        for w in list(set(article)):
            if w not in word_file_count:
                word_file_count[w] = 0
            word_file_count[w] += 1
    return word_file_count

# 计算TF-IDF,并返回字典
def tf_idf(word_list, word_file_count, entity_list):
    out_dict = {}
    freq_word_dict = freq_word(word_list)
    for entity in entity_list:
        # 计算TF：某个词在文章中出现的次数/文章总词数
        tf = freq_word_dict[str(entity)] / len(word_list)
        # 计算IDF：log(语料库的文档总数/(包含该词的文档数+1))
        idf = math.log(word_file_count["total article num"] / (word_file_count[str(entity)] + 1), 10)
        # 计算TF-IDF
        tfidf = tf * idf * 100
        out_dict[str(entity)] = float('%.4f'% tfidf)
    return out_dict


datafile = 'news_copy.xlsx'
wb = load_workbook(datafile)
ws = wb.active
ws.cell(row=1, column=4).value = 'key_words'

# 所有的文章
article_list = []
# 识别出的实体列表
entity_list = []
noun_flag = ['n', 'nr', 'nr1', 'nr2', 'nrj', 'nrf', 'ns', 'nsf', 'nt', 'nz', 'nl', 'ng']

for i in tqdm(range(2, ws.max_row+1)):
    article = ws.cell(row=i, column=2).value
    entity = []
    seg_article, flag_article = zip(*[(w.word, w.flag) for w in pseg.cut(article)])
    for a in range(len(flag_article)):
        if flag_article[a] in noun_flag:
            entity.append(seg_article[a])
    entity_list.append(entity)
    article_list.append(list(seg_article))

word_file_count = word_in_file_count(article_list)
with open('word_file_count.json', 'w', encoding='utf-8') as f:
    json.dump(word_file_count, f)

entity_count = {}

for i in tqdm(range(len(article_list))):
    result = tf_idf(article_list[i], word_file_count, entity_list[i])
    for entity in result:
        if  entity not in entity_count.keys():
            entity_count[entity] = {"num": 0, "tfidf": []}
        entity_count[entity]["num"] += 1
        entity_count[entity]["tfidf"].append(result[entity])
    order_result = sorted(result.items(), key=operator.itemgetter(1), reverse=True)
    ws.cell(row=i+2, column=4).value = json.dumps(order_result[:20], ensure_ascii=False)

with open('entity_count.txt', 'w', encoding='utf-8') as f:
    for entity in entity_count.keys():
        num = entity_count[entity]["num"]
        tfidf_list = entity_count[entity]["tfidf"]
        f.write("%s\t%d\t%.6f\t%.6f\n" % (entity, num, sum(tfidf_list), sum(tfidf_list)/len(tfidf_list)))

wb.save(datafile)

#word_file_count = json.load(open('word_file_count.json'))
