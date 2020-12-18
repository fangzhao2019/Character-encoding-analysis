from openpyxl import load_workbook
import os

path='cluster_results'
recogResult={}
wordNames=os.listdir(path)
wordNames.remove('null')
a=0
for word in wordNames:
    imageNames=os.listdir('%s/%s'%(path,word))
    a+=len(imageNames)
    for imageName in imageNames:
        ttf_name,image=imageName.split('_uni')
        code=image.replace('.png','')
        if not ttf_name in recogResult.keys():
            recogResult[ttf_name]={}
        recogResult[ttf_name][code]=word

wb=load_workbook('coding3.xlsx')
ws=wb.active

for i in range(2,ws.max_row+1):
    
    ttf_name=ws.cell(row=i,column=2).value
    code=ws.cell(row=i,column=3).value
    word=ws.cell(row=i,column=4).value

    try:
        new_word=recogResult[ttf_name][code]
        ws.cell(row=i,column=7).value=new_word
    except:
        ws.cell(row=i,column=7).value='CHANGE'
    if i%1000==0:
        print('已处理数据%d条'%i)
    i+=1
wb.save('finalCoding.xlsx')
