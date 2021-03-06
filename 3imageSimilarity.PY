from PIL import Image
import numpy as np
import random
import os
import shutil
from openpyxl import load_workbook
import time

def dataSetLoading(path):
    dataSet={}
    wordNames=os.listdir(path)
    wordNames.remove('null') 
    for word in wordNames:
        print('正在载入<%s>里的数据'%word)
        imageNames=os.listdir('%s/%s'%(path,word))
        random_imageNames=random.sample(imageNames,5)
        imageSet=[]
        for imageName in random_imageNames:
            imagePath='%s/%s/%s'%(path,word,imageName)
            image=np.array(Image.open(imagePath))[:,:,0]
            imageSet.append(image)
        dataSet[word]=np.array(imageSet)
    return dataSet

def similarityMeasure(image,dataSet):
    best_sim=0
    best_key='null'
    imageExtend=np.tile(image,(5,1,1))
    for key in dataSet.keys():
        imageSet=dataSet[key]
        sim=sum(sum(sum(imageExtend==imageSet)))/(2048*2048*5)
        if sim>0.9:
            return key
        if sim>best_sim:
            best_sim=sim
            best_key=key
    if best_sim<0.5:
        best_key='null'
    return best_key

time1=time.time()
dataSet=dataSetLoading('cluster_results')

print('正在处理未识别的图片')
wb=load_workbook('coding2.xlsx')
ws=wb.active

for i in range(2,ws.max_row+1):
    ttf_name=ws.cell(row=i,column=2).value
    code=ws.cell(row=i,column=3).value
    word=ws.cell(row=i,column=4).value
    if not word=='null':continue
    if i%10==0:
        print('已处理数据%d条'%i)
    i+=1

    imagePath='cluster_results/null/%s_uni%s.png'%(ttf_name,code)
    image=np.array(Image.open(imagePath))[:,:,0]
    recogResult=similarityMeasure(image,dataSet)

    ws.cell(row=i,column=6).value=recogResult

    if not os.path.exists('recogResult/%s'%recogResult):
        os.mkdir('recogResult/%s'%recogResult)
    newPath='recogResult/%s/%s_uni%s.png'%(recogResult,ttf_name,code)
    shutil.copyfile(imagePath,newPath)

wb.save('coding3.xlsx')
    
time2=time.time()
print('共耗时%d秒'%(time2-time1))

    





        
    
