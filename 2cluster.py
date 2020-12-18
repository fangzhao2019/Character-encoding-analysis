from openpyxl import load_workbook
import os
import shutil

wb=load_workbook('init_coding.xlsx')
ws=wb.active

changeNum=0

for i in range(2,ws.max_row+1):
    if i%1000==0:print(i)
    i+=1
    ttf_name=ws.cell(row=i,column=2).value
    code=ws.cell(row=i,column=3).value
    word=ws.cell(row=i,column=4).value

    folder='cluster_results/%s'%word
    if not os.path.exists(folder):
        os.mkdir(folder)

    old_filepath='ttf_images/ttf_images_cache/%s/uni%s.png'%(ttf_name,code)
    new_filepath='%s/%s_uni%s.png'%(folder,ttf_name,code)
    try:
        shutil.copyfile(old_filepath,new_filepath)
    except:
        ws.cell(row=i,column=5).value='CHANGE'
        changeNum+=1

wb.save('coding2.xlsx')

    
        
        
